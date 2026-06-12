from datetime import datetime
from decimal import Decimal, InvalidOperation
from threading import Thread
from typing import Any

from django.conf import settings
import openpyxl
from rest_framework.viewsets import ViewSet
from rest_framework.request import Request
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from rest_framework.status import HTTP_201_CREATED
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count
from django.contrib.auth import get_user_model

from ..volulink.utils import send_html_email
from ..volulink.responses import ErrorWithMessageResponse, SuccessWithMessageResponse, SuccessWithResultsResponse
from ..authentication.constants import ROLE_ADMIN, ROLE_FACILITY_MANAGER, ROLE_FEDERATION_MANAGER
from ..authentication.permissions import IsAdmin, IsFacilityManager, IsFederationManager, MultiRolePermission
from .models import Category, Facility, FacilityDetail, FacilityActivityLog
from .serializers import (
    CategorySerializer,
    FacilityRestrictedSerializer,
    FederationSerializer,
    InternalBenchmarkQueryParamsSerializer,
    UnapprovedUserFacilitySerializer,
    FacilityDetailRetrieveSerializer,
    FacilityDetailSerializer,
    FacilitySerializer,
    FacilityWithDetailSerializer,
    BenchmarkQueryParamsSerializer,
)
from .benchmark import CategoryWideBenchmark, InternalBenchmark

User = get_user_model()


class CategoryApi(ViewSet):
    def get_permissions(self):
        return [MultiRolePermission(ROLE_ADMIN, ROLE_FACILITY_MANAGER, ROLE_FEDERATION_MANAGER)] if self.action == 'list' else [IsAdmin()]
    
    @staticmethod
    def list(request: Request) -> SuccessWithResultsResponse:
        queryset = Category.objects.all()
        serializer = CategorySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    def create(request: Request) -> SuccessWithMessageResponse:
        serializer = CategorySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return SuccessWithMessageResponse('Es wurde erfolgreich eine Kategorie erstellt.', status=HTTP_201_CREATED)

    @staticmethod
    def retrieve(request: Request, pk) -> SuccessWithMessageResponse:
        serializer = CategorySerializer(get_object_or_404(Category, pk=pk))
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    def update(request: Request, pk: int) -> SuccessWithMessageResponse:
        category = get_object_or_404(Category, pk=pk)
        serializer = CategorySerializer(category, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return SuccessWithMessageResponse('Die Kategorie wurde erfolgreich aktualisiert.')
    
    @staticmethod
    def destroy(request: Request, pk: int) -> SuccessWithMessageResponse:
        category = get_object_or_404(Category, pk=pk)
        category.delete()
        return SuccessWithMessageResponse('Die Kategorie wurde erfolgreich gelöscht.')


class UnapprovedUserFacilityApi(ViewSet):
    permission_classes = [AllowAny]

    @staticmethod
    def list(request: Request) -> SuccessWithResultsResponse:
        queryset = Facility.objects.filter(user__isnull=True, is_federation=False)
        serializer = UnapprovedUserFacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)


class UnApprovedFedarationApi(ViewSet):
    permission_classes = [AllowAny]

    @staticmethod
    def list(request: Request) -> SuccessWithResultsResponse:
        queryset = Facility.objects.filter(user__isnull=True, is_federation=True)
        serializer = UnapprovedUserFacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)


class FederationApi(ViewSet):
    permission_classes = [
        IsAdmin |
        IsFederationManager
    ]

    @staticmethod
    def list(request: Request) -> SuccessWithResultsResponse:
        if request.user.role == ROLE_ADMIN:
            queryset = Facility.get_federations()
        else:
            queryset = Facility.get_own_federations(request.user)
        queryset = queryset.annotate(
            facility_count=Count('facilities')
        )
        serializer = FederationSerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)

    @staticmethod
    @action(detail=True, methods=['get'])
    def facilities(request: Request, pk: int) -> SuccessWithResultsResponse:
        if request.user.role == ROLE_ADMIN:
            queryset = Facility.get_facilities_by_federation(pk)
        else:
            queryset = Facility.get_facilities_by_own_federation(pk, request.user)
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)


class FacilityApi(ViewSet):
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'update']:
            return [MultiRolePermission(ROLE_ADMIN, ROLE_FACILITY_MANAGER, ROLE_FEDERATION_MANAGER)]
        elif self.action in [
            'facility_detail',
            'retrieve_detail',
            'update_unpublished_detail',
            'benchmark'
        ]:
            return [MultiRolePermission(ROLE_FACILITY_MANAGER, ROLE_FEDERATION_MANAGER)]
        elif self.action == 'facilities_of_specific_federation':
            return [IsFederationManager()]
        return [IsAdmin()]

    def list(self, request: Request) -> SuccessWithResultsResponse:
        if request.user.role == ROLE_ADMIN:
            return self.facility_list_by_admin()
        elif request.user.role == ROLE_FACILITY_MANAGER:
            return self.facility_list_by_facility_manager(request)
        return self.facility_list_by_federation_manager(request)
    
    @staticmethod
    def facility_list_by_admin() -> SuccessWithResultsResponse:
        queryset = Facility.objects.select_related('category').all()
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    def facility_list_by_facility_manager(request: Request) -> SuccessWithResultsResponse:
        queryset = (
            Facility.objects
            .select_related('category')
            .filter(
                user=request.user,
                is_user_approved=True
            )
        )
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    def facility_list_by_federation_manager(request: Request) -> SuccessWithResultsResponse:
        queryset = (
            Facility.objects
            .select_related('category')
            .filter(
                user=request.user,
                is_user_approved=True
            )
        )
        federation_ids = [item.id for item in queryset if item.is_federation]
        facilities_in_federations = (
            Facility.objects
            .select_related('category')
            .filter(federation_id__in=federation_ids)
        )
        queryset = queryset.union(facilities_in_federations)
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)

    def create(self, request: Request) -> SuccessWithMessageResponse:
        serializer = FacilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        facility = serializer.save()

        thread = Thread(
            target=self.send_facility_create_email,
            args=(facility,)
        )
        thread.daemon = True
        thread.start()

        return SuccessWithMessageResponse('Eine Anlage wurde erfolgreich errichtet.', status=HTTP_201_CREATED)
    
    @staticmethod
    def send_facility_create_email(facility: Facility) -> None:
        admin_users = User.objects.filter(role=ROLE_ADMIN).values_list('email', flat=True)
        facility_type = 'Einrichtung' if not facility.is_federation else 'Verbund'
        context = {
            'facility_name': facility.name,
            'facility_type': facility_type,
            'administration_link': settings.APP_LINK + '/facilities/' + facility.id + '/update'
        }

        send_html_email(
            subject='Eine neue ' + facility_type + ' wurde geschaffen | Benchmarking-Tool',
            template_name='emails/new_facility.html',
            context=context,
            recipient_list=admin_users
        )
    
    def retrieve(self, request: Request, pk: int) -> SuccessWithResultsResponse:
        if request.user.role == ROLE_ADMIN:
            return self.retrieve_by_admin(request, pk)
        return self.retrieve_by_facility_or_federation_manager(request, pk)
    
    @staticmethod
    def retrieve_by_facility_or_federation_manager(request: Request, pk: int) -> SuccessWithResultsResponse:
        facility = get_object_or_404(
            Facility.objects.select_related('category'),
            Q(
                (
                    Q(
                        Q(federation__user=request.user) | Q(user=request.user),
                        federation__isnull=False
                    ) | Q(
                        federation__isnull=True,
                        user=request.user
                    )
                ),
                pk=pk,
                is_user_approved=True
            )
        )
        serializer = FacilityWithDetailSerializer(facility, context={'facility': facility})
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    def retrieve_by_admin(request: Request, pk: int) -> SuccessWithResultsResponse:
        facility = get_object_or_404(Facility.objects.select_related('category'), pk=pk)
        serializer = FacilitySerializer(facility)
        return SuccessWithResultsResponse(serializer.data)

    @staticmethod
    @action(methods=['get'], detail=True, url_path='of-federation')
    def facilities_of_specific_federation(request: Request, pk: int) -> SuccessWithResultsResponse:
        queryset = (
            Facility.objects
            .select_related('category')
            .filter(
                federation_id=pk
            )
        )
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)

    @staticmethod
    def update(request: Request, pk: int) -> SuccessWithMessageResponse:
        facility = get_object_or_404(Facility, pk=pk)

        if request.user.role == IsAdmin:
            serializer = FacilitySerializer(facility, data=request.data, partial=True)
        else:
            serializer = FacilityRestrictedSerializer(facility, data=request.data, partial=True)

        serializer.is_valid(raise_exception=True)
        serializer.save()
        return SuccessWithMessageResponse('Die Anlage wurde erfolgreich aktualisiert.')
    
    @staticmethod
    @action(methods=['post'], detail=True, url_path='detail')
    def facility_detail(request: Request, pk: int) -> SuccessWithMessageResponse:
        facility = get_object_or_404(
            Facility,
            pk=pk,
            user=request.user,
            is_user_approved=True
        )
        data = request.data.copy()
        data['facility'] = facility.id
        serializer = FacilityDetailSerializer(data=data, context={'facility': facility})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return SuccessWithMessageResponse('The facility\'s detail has been successfully updated.')

    @action(methods=['get'], detail=True, url_path=r'detail/(?P<facility_detail_pk>[^/.]+)')
    def retrieve_detail(self, request: Request, pk: int, facility_detail_pk: int) -> SuccessWithResultsResponse:
        facility_detail = get_object_or_404(
            FacilityDetail.objects.select_related('facility__federation', 'facility__category'),
            pk=facility_detail_pk,
            facility_id=pk,
            facility__user=request.user,
            facility__is_user_approved=True
        )
        serializer = FacilityDetailRetrieveSerializer(
            facility_detail,
            context={'facility': facility_detail.facility}
        )
        return SuccessWithResultsResponse(serializer.data)

    @staticmethod
    @retrieve_detail.mapping.put
    def update_unpublished_detail(request: Request, pk: int, facility_detail_pk: int) -> SuccessWithMessageResponse:
        facility_detail = get_object_or_404(
            FacilityDetail.objects.select_related('facility__category'),
            pk=facility_detail_pk,
            facility_id=pk,
            facility__user=request.user,
            facility__is_user_approved=True
        )
        serializer = FacilityDetailSerializer(
            facility_detail,
            data=request.data,
            partial=True,
            context={'facility': facility_detail.facility}
        )
        serializer.is_valid(raise_exception=True)

        if serializer.validated_data['is_published']:
            serializer.save(last_published_at=datetime.now())
        else:
            serializer.save()
        return SuccessWithMessageResponse('The facility\'s detail has been successfully updated.')
    
    @staticmethod
    @action(methods=['get'], detail=False, url_path='user-joining-requests')
    def user_joining_request_list(request: Request) -> SuccessWithResultsResponse:
        queryset = (
            Facility.objects
            .select_related('category')
            .filter(
                is_user_approved=False,
                user_id__isnull=False
            )
        )
        serializer = FacilitySerializer(queryset, many=True)
        return SuccessWithResultsResponse(serializer.data)
    
    @staticmethod
    @action(methods=['put'], detail=True, url_path='approve-users')
    def approve_user(request: Request, pk: int) -> SuccessWithMessageResponse:
        facility = get_object_or_404(Facility, pk=pk, is_user_approved=False)
        facility.is_user_approved = True
        facility.save()
        FacilityActivityLog.log_facility_user_approve(facility, request.user)
        return SuccessWithMessageResponse('Der Benutzer wurde für die Einrichtung zugelassen.')
    
    @staticmethod
    @action(methods=['delete'], detail=True, url_path='detach-users')
    def detach_user(request: Request, pk: int) -> SuccessWithMessageResponse:
        facility = get_object_or_404(Facility, pk=pk)
        facility.is_user_approved = False
        facility.save()
        facility.user.delete()
        FacilityActivityLog.log_facility_user_reject(facility, request.user)
        return SuccessWithMessageResponse('Der Benutzer wurde von der Anlage getrennt.')
    
    @staticmethod
    def destroy(request: Request, pk: int) -> SuccessWithMessageResponse:
        facility = get_object_or_404(Facility, pk=pk)
        facility.delete()
        return SuccessWithMessageResponse('Die Einrichtung wurde erfolgreich gelöscht.')


class InternalBenchmarkApi(ViewSet):
    permission_classes = [IsFederationManager]

    @staticmethod
    def list(request: Request) -> SuccessWithResultsResponse:
        federation_id = request.query_params.get('federation', None)

        if not federation_id:
            return ErrorWithMessageResponse('Eine Föderations-ID ist erforderlich.')

        years = (
            FacilityDetail.objects
            .filter(
                is_published=True,
                facility__federation_id=federation_id,
                facility__federation__user=request.user,
                facility__federation__is_user_approved=True
            )
            .values_list('year', flat=True)
            .distinct()
        )
        return SuccessWithResultsResponse(years)
    
    @staticmethod
    @action(detail=False, methods=['get'], url_path='benchmark')
    def benchmark(request: Request) -> SuccessWithResultsResponse:
        serializer = InternalBenchmarkQueryParamsSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        year = serializer.validated_data['year']
        federation_id = serializer.validated_data['federation']

        query = get_object_or_404(
            Facility,
            pk=federation_id,
            is_federation=True,
            user=request.user,
            is_user_approved=True
        )
        benchmark = InternalBenchmark(query, year).build()
        FacilityActivityLog.log_benchmark_attempt(query, request.user, year)
        return SuccessWithResultsResponse(benchmark)


class CategoryWideBenchmarkApi(ViewSet):
    permission_classes = [
        IsFederationManager |
        IsFacilityManager
    ]

    @staticmethod
    @action(detail=False, methods=['get'], url_path='eligible-facilities')
    def eligible_facilities(request: Request) -> SuccessWithResultsResponse:
        queryset = (
            Facility.objects
            .select_related('category')
            .prefetch_related('facility_details')
            .filter(
                Q(
                    user=request.user,
                    federation__isnull=True
                ) |
                Q(
                    federation__isnull=False,
                    federation__user=request.user
                )
            )
            .exclude(
                facility_details__isnull=True
            )
        )
        results = [{
            'id': item.id,
            'name': item.name,
            'category': item.category_id,
            'category_name': item.category.name,
            'years': [detail.year for detail in item.facility_details.all() if detail.is_published]
        } for item in queryset]
        return SuccessWithResultsResponse(results)

    @staticmethod
    @action(detail=False, methods=['get'])
    def benchmark(request: Request) -> SuccessWithResultsResponse:
        serializer = BenchmarkQueryParamsSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        facility_id = serializer.validated_data['facility']
        year = int(serializer.validated_data['year'])
        facility = get_object_or_404(
            Facility.objects.select_related('category'),
            Q(
                user=request.user,
                federation__isnull=True
            ) |
            Q(
                federation__isnull=False,
                federation__user=request.user
            ),
            pk=facility_id,
            facility_details__year=year
        )

        benchmark_eligibility_check = (
            FacilityDetail.objects
            .filter(
                year=year,
                is_published=True,
                facility__category_id=facility.category_id
            )
            .count()
        )
        if benchmark_eligibility_check < 5:
            return ErrorWithMessageResponse(
                'Für einen Vergleich in der ausgewählten Kategorie liegen nicht genügend Daten vor.'
            )

        return SuccessWithResultsResponse({
            'facility': {
                'id': facility_id,
                'name': facility.name,
                'category': facility.category_id,
                'category_name': facility.category.name
            },
            'benchmark': CategoryWideBenchmark(facility, year).build()
        })


class MasterDataApi(ViewSet):
    permission_classes = [IsAdmin]

    @staticmethod
    def safe_int(value: Any) -> int | None:
        if value is None or value != value:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def safe_decimal(value: Any) -> Decimal | None:
        if value is None or value != value:
            return None
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError):
            return None

    @staticmethod
    def get_cell(sheet: list[list], row: int, col: int) -> Any:
        try:
            val = sheet[row][col]
            if val is None or val != val:
                return None
            return val
        except IndexError:
            return None

    @staticmethod
    def load_sheet(file, sheet_name: str) -> list[list]:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        return [[cell.value for cell in row] for row in ws.iter_rows()]

    def parse_sheet(self, sheet: list[list]) -> list[dict]:
        ROW = {
            'name': 0,
            'category_id': 1,
            'rooms': 3,
            'beds': 3,
            'opening_days_per_year': 14,
            'year': 5,
            'guests': 7,
            'rooms_sold': 18,
            'overnight_stays': 25,
            'total_revenue': 28,
            'donations_subsidies_income': 29,
            'personnel_costs': 33,
            'material_goods_costs': 32,
            'energy_costs': 34,
            'other_operating_costs': 35,
        }

        FIRST_DATA_COL = 1
        facilities = []
        total_cols = max(len(row) for row in sheet)

        col = FIRST_DATA_COL
        while col < total_cols:
            name = self.get_cell(sheet, ROW['name'], col)
            if not isinstance(name, str) or not name.strip():
                col += 2
                continue

            facilities.append({
                'facility': {
                    'name': name.strip(),
                    'category_id': self.safe_int(self.get_cell(sheet, ROW['category_id'], col + 1)),
                    'rooms': self.safe_int(self.get_cell(sheet, ROW['rooms'], col)),
                    'beds': self.safe_int(self.get_cell(sheet, ROW['beds'], col + 1)),
                    'opening_days_per_year': self.safe_int(self.get_cell(sheet, ROW['opening_days_per_year'], col)) or 365,
                },
                'detail': {
                    'year': self.safe_int(self.get_cell(sheet, ROW['year'], col)),
                    'guests': self.safe_int(self.get_cell(sheet, ROW['guests'], col)) or 0,
                    'rooms_sold': self.safe_int(self.get_cell(sheet, ROW['rooms_sold'], col)) or 0,
                    'overnight_stays': self.safe_int(self.get_cell(sheet, ROW['overnight_stays'], col)) or 0,
                    'total_revenue': self.safe_decimal(self.get_cell(sheet, ROW['total_revenue'], col)) or Decimal('0.00'),
                    'donations_subsidies_income': self.safe_decimal(self.get_cell(sheet, ROW['donations_subsidies_income'], col)),
                    'personnel_costs': self.safe_decimal(self.get_cell(sheet, ROW['personnel_costs'], col)) or Decimal('0.00'),
                    'material_goods_costs': self.safe_decimal(self.get_cell(sheet, ROW['material_goods_costs'], col)) or Decimal('0.00'),
                    'energy_costs': self.safe_decimal(self.get_cell(sheet, ROW['energy_costs'], col)) or Decimal('0.00'),
                    'outsourced_services_costs': Decimal('0.00'),
                    'other_operating_costs': self.safe_decimal(self.get_cell(sheet, ROW['other_operating_costs'], col)) or Decimal('0.00'),
                },
            })
            col += 2

        return facilities

    def create(self, request: Request) -> ErrorWithMessageResponse | SuccessWithResultsResponse:
        file = request.FILES.get('file')
        if not file:
            return ErrorWithMessageResponse('Keine Datei bereitgestellt.')
        if not file.name.endswith('.xlsx'):
            return ErrorWithMessageResponse('Es werden nur .xlsx-Dateien unterstützt.')
        
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet_names = wb.sheetnames
        except Exception as e:
            return ErrorWithMessageResponse(f'Datei konnte nicht gelesen werden: {e}')
        
        results = {'created': 0, 'updated': 0, 'errors': []}

        for sheet_name in sheet_names:
            sheet = [[cell.value for cell in row] for row in wb[sheet_name].iter_rows()]
            entries = self.parse_sheet(sheet)

            for entry in entries:
                facility_data = entry['facility']
                detail_data = entry['detail']

                try:
                    facility, created = Facility.objects.update_or_create(
                        name=facility_data['name'],
                        defaults={
                            'category_id': facility_data['category_id'],
                            'rooms': facility_data['rooms'],
                            'beds':  facility_data['beds'],
                            'opening_days_per_year': facility_data['opening_days_per_year'],
                        }
                    )

                    FacilityDetail.objects.update_or_create(
                        facility=facility,
                        year=detail_data['year'],
                        is_published=True,
                        defaults={k: v for k, v in detail_data.items() if k != 'year'}
                    )

                    if created:
                        results['created'] += 1
                    else:
                        results['updated'] += 1

                except Exception as e:
                    results['errors'].append({'facility': facility_data['name'], 'error': str(e)})

        return SuccessWithResultsResponse(results)
